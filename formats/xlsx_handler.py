"""
XLSX/XLS format handler — spreadsheet extraction and reconstruction via openpyxl.

Ingest:
  Each sheet → H2 section + TABLE element.
  Formulas captured from non-data_only workbook.
  Merged cells unmerged with duplicated values.
  .xls files are first converted to .xlsx via LibreOffice headless.

Export:
  Rebuilds workbook from H2-delimited TABLE elements.
  Tier 2: restores column widths, number formats, formulas, cell styles from sidecar.
  Tier 3: if original .xlsx exists and hash match ≥ 80%, patches cell values in-place.
"""

import time
from pathlib import Path
from typing import Any

import structlog

from formats.base import FormatHandler, register_handler
from core.document_model import (
    DocumentModel,
    DocumentMetadata,
    Element,
    ElementType,
    compute_content_hash,
)

log = structlog.get_logger(__name__)


@register_handler
class XlsxHandler(FormatHandler):
    EXTENSIONS = ["xlsx", "xls"]

    # ── Ingest ────────────────────────────────────────────────────────────────

    def ingest(self, file_path: Path) -> DocumentModel:
        import openpyxl

        file_path = Path(file_path)
        t_start = time.perf_counter()
        log.info("handler_ingest_start", filename=file_path.name, format="xlsx")
        _tmp_xlsx: Path | None = None

        if file_path.suffix.lower() == ".xls":
            from core.libreoffice_helper import convert_with_libreoffice

            file_path = convert_with_libreoffice(file_path, "xlsx")
            _tmp_xlsx = file_path

        try:
            model = DocumentModel()
            model.metadata = DocumentMetadata(
                source_file=file_path.name,
                source_format="xlsx",
            )

            # Open with data_only=True to get computed values
            wb_data = openpyxl.load_workbook(str(file_path), data_only=True)
            # Also open without data_only to capture formulas
            try:
                wb_formula = openpyxl.load_workbook(str(file_path), data_only=False)
            except Exception:
                wb_formula = None

            sheet_count = len(wb_data.sheetnames)
            model.metadata.page_count = sheet_count

            for sheet_idx, sheet_name in enumerate(wb_data.sheetnames):
                ws_data = wb_data[sheet_name]
                ws_formula = wb_formula[sheet_name] if wb_formula else None

                # Add separator between sheets (not before first)
                if sheet_idx > 0:
                    model.add_element(Element(type=ElementType.HORIZONTAL_RULE, content=""))

                # Sheet heading
                model.add_element(
                    Element(type=ElementType.HEADING, content=sheet_name, level=2)
                )

                # Handle merged cells — unmerge and duplicate values
                merged_cells_map = self._build_merged_cells_map(ws_data)

                # Build table from used range
                if ws_data.max_row is None or ws_data.max_column is None:
                    continue
                if ws_data.max_row < 1 or ws_data.max_column < 1:
                    continue

                rows: list[list[str]] = []
                for row_idx in range(ws_data.min_row, ws_data.max_row + 1):
                    row_data: list[str] = []
                    for col_idx in range(ws_data.min_column, ws_data.max_column + 1):
                        cell = ws_data.cell(row=row_idx, column=col_idx)
                        value = cell.value

                        # Check merged cell map for duplicated values
                        if value is None and (row_idx, col_idx) in merged_cells_map:
                            value = merged_cells_map[(row_idx, col_idx)]

                        row_data.append(self._format_cell_value(value))
                    rows.append(row_data)

                # Trim trailing empty rows
                while rows and all(c == "" for c in rows[-1]):
                    rows.pop()

                # Trim trailing empty columns
                if rows:
                    while rows[0] and all(row[-1] == "" for row in rows if row):
                        for row in rows:
                            if row:
                                row.pop()

                if rows:
                    model.add_element(Element(type=ElementType.TABLE, content=rows))

                # Extract images
                self._extract_sheet_images(ws_data, model)

            wb_data.close()
            if wb_formula:
                wb_formula.close()

            duration_ms = int((time.perf_counter() - t_start) * 1000)
            log.info(
                "handler_ingest_complete",
                filename=file_path.name,
                element_count=len(model.elements),
                duration_ms=duration_ms,
            )
            return model
        finally:
            if _tmp_xlsx and _tmp_xlsx.exists():
                _tmp_xlsx.unlink(missing_ok=True)

    def _build_merged_cells_map(self, ws: Any) -> dict[tuple[int, int], Any]:
        """Build a map of merged cell coordinates to the merge's top-left value."""
        merged_map: dict[tuple[int, int], Any] = {}
        for merge_range in list(ws.merged_cells.ranges):
            min_row, min_col, max_row, max_col = (
                merge_range.min_row,
                merge_range.min_col,
                merge_range.max_row,
                merge_range.max_col,
            )
            # Get the value from the top-left cell
            top_left_value = ws.cell(row=min_row, column=min_col).value
            # Map all other cells in the range to this value
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if (r, c) != (min_row, min_col):
                        merged_map[(r, c)] = top_left_value
        return merged_map

    def _format_cell_value(self, value: Any) -> str:
        """Format a cell value to string, preserving type hints."""
        if value is None:
            return ""
        if isinstance(value, float):
            if value == int(value):
                return str(int(value))
            return str(value)
        return str(value)

    def _extract_sheet_images(self, ws: Any, model: DocumentModel) -> None:
        from core.image_handler import extract_image
        from core.document_model import ImageData

        try:
            for img in ws._images:
                raw = img._data()
                if raw:
                    hash_name, png_data, meta = extract_image(raw, "png")
                    model.images[hash_name] = ImageData(
                        data=png_data,
                        original_format="xlsx_embedded",
                        width=meta.get("width"),
                        height=meta.get("height"),
                    )
                    model.add_element(
                        Element(
                            type=ElementType.IMAGE,
                            content=f"assets/{hash_name}",
                        )
                    )
        except Exception as exc:
            log.debug("xlsx.image_extract_failed", error=str(exc))

    # ── Export ─────────────────────────────────────────────────────────────────

    def export(
        self,
        model: DocumentModel,
        output_path: Path,
        sidecar: dict[str, Any] | None = None,
        original_path: Path | None = None,
    ) -> None:
        import openpyxl

        tier = 3 if (original_path and original_path.exists() and sidecar) else (2 if sidecar else 1)
        t_start = time.perf_counter()
        log.info("handler_export_start", filename=output_path.name, target_format="xlsx", tier=tier)

        # Tier 3: patch original
        if original_path and original_path.exists() and sidecar:
            if self._try_tier3_export(model, output_path, original_path, sidecar):
                duration_ms = int((time.perf_counter() - t_start) * 1000)
                log.info("handler_export_complete", filename=output_path.name, output_path=str(output_path), duration_ms=duration_ms)
                return

        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active

        sections = self._split_into_sheet_sections(model)

        for idx, (sheet_name, elements) in enumerate(sections):
            if idx == 0 and default_sheet:
                ws = default_sheet
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)

            self._write_sheet(ws, elements, sidecar, sheet_name)

        # Remove default sheet if no sections used it
        if not sections and default_sheet:
            wb.remove(default_sheet)

        wb.save(str(output_path))
        duration_ms = int((time.perf_counter() - t_start) * 1000)
        log.info("handler_export_complete", filename=output_path.name, output_path=str(output_path), duration_ms=duration_ms)

    def _split_into_sheet_sections(
        self, model: DocumentModel
    ) -> list[tuple[str, list[Element]]]:
        """Split model into (sheet_name, elements) tuples at H2 boundaries."""
        sections: list[tuple[str, list[Element]]] = []
        current_name = "Sheet1"
        current_elements: list[Element] = []

        for elem in model.elements:
            if elem.type == ElementType.HORIZONTAL_RULE:
                if current_elements:
                    sections.append((current_name, current_elements))
                    current_elements = []
                    current_name = f"Sheet{len(sections) + 2}"
                continue

            if elem.type == ElementType.HEADING and elem.level == 2:
                if current_elements:
                    sections.append((current_name, current_elements))
                    current_elements = []
                current_name = str(elem.content)[:31]  # Excel sheet name limit
                continue

            current_elements.append(elem)

        if current_elements:
            sections.append((current_name, current_elements))

        return sections

    def _write_sheet(
        self,
        ws: Any,
        elements: list[Element],
        sidecar: dict[str, Any] | None,
        sheet_name: str,
    ) -> None:
        """Write elements to a worksheet."""
        import openpyxl
        from openpyxl.utils import get_column_letter

        for elem in elements:
            if elem.type != ElementType.TABLE:
                continue

            rows = elem.content
            if not isinstance(rows, list):
                continue

            for ri, row in enumerate(rows, start=1):
                for ci, cell_value in enumerate(row, start=1):
                    cell = ws.cell(row=ri, column=ci)
                    cell.value = self._parse_cell_value(cell_value)

            # Apply sidecar styles
            if sidecar:
                elements_data = sidecar.get("elements", {})
                h = compute_content_hash(rows)
                style_info = elements_data.get(h, {})

                # Column widths
                col_widths = style_info.get("column_widths", {})
                for col_letter, width in col_widths.items():
                    ws.column_dimensions[col_letter].width = width

                # Number formats
                number_formats = style_info.get("number_formats", {})
                for cell_ref, fmt in number_formats.items():
                    try:
                        ws[cell_ref].number_format = fmt
                    except Exception:
                        pass

                # Formulas
                formulas = style_info.get("formulas", {})
                for cell_ref, formula in formulas.items():
                    try:
                        ws[cell_ref].value = formula
                    except Exception:
                        pass

                # Merged cells
                merged_ranges = style_info.get("merged_cells", [])
                for merge_range in merged_ranges:
                    try:
                        ws.merge_cells(merge_range)
                    except Exception:
                        pass

                # Cell styles (font, fill, alignment)
                cell_styles = style_info.get("cell_styles", {})
                for cell_ref, cell_style in cell_styles.items():
                    try:
                        cell = ws[cell_ref]
                        self._apply_cell_style(cell, cell_style)
                    except Exception:
                        pass

            break  # One table per sheet

    def _parse_cell_value(self, text: str) -> Any:
        """Parse string cell value back to appropriate type."""
        if not text:
            return None
        # Try int
        try:
            return int(text)
        except ValueError:
            pass
        # Try float
        try:
            return float(text)
        except ValueError:
            pass
        return text

    def _apply_cell_style(self, cell: Any, style: dict) -> None:
        """Apply cell style from sidecar."""
        from openpyxl.styles import Font, PatternFill, Alignment

        font_info = style.get("font", {})
        if font_info:
            cell.font = Font(
                name=font_info.get("name"),
                size=font_info.get("size"),
                bold=font_info.get("bold", False),
                italic=font_info.get("italic", False),
            )

        fill_info = style.get("fill", {})
        if fill_info and fill_info.get("color"):
            cell.fill = PatternFill(
                start_color=fill_info["color"],
                end_color=fill_info["color"],
                fill_type="solid",
            )

        align_info = style.get("alignment", {})
        if align_info:
            cell.alignment = Alignment(
                horizontal=align_info.get("horizontal", "general"),
                vertical=align_info.get("vertical", "bottom"),
                wrap_text=align_info.get("wrap_text", False),
            )

    def _try_tier3_export(
        self,
        model: DocumentModel,
        output_path: Path,
        original_path: Path,
        sidecar: dict[str, Any],
    ) -> bool:
        """Attempt Tier 3: patch cell values in original XLSX."""
        import openpyxl

        try:
            wb = openpyxl.load_workbook(str(original_path))
        except Exception:
            return False

        sections = self._split_into_sheet_sections(model)

        # Check hash match
        original_values: list[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        original_values.append(str(cell.value))

        if not original_values:
            wb.close()
            return False

        model_values: set[str] = set()
        for _, elems in sections:
            for elem in elems:
                if elem.type == ElementType.TABLE and isinstance(elem.content, list):
                    for row in elem.content:
                        for cell in row:
                            if cell:
                                model_values.add(str(cell))

        match_count = sum(1 for v in original_values if v in model_values)
        match_ratio = match_count / len(original_values) if original_values else 0

        if match_ratio < 0.8:
            wb.close()
            return False

        # Patch values
        for sheet_name, elems in sections:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for elem in elems:
                    if elem.type == ElementType.TABLE and isinstance(elem.content, list):
                        for ri, row in enumerate(elem.content, start=1):
                            for ci, val in enumerate(row, start=1):
                                try:
                                    cell = ws.cell(row=ri, column=ci)
                                    # Only update data cells, not formulas
                                    if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                                        cell.value = self._parse_cell_value(val)
                                except AttributeError:
                                    # MergedCell objects have read-only .value
                                    pass

        wb.save(str(output_path))
        wb.close()
        return True

    # ── Style extraction ──────────────────────────────────────────────────────

    def extract_styles(self, file_path: Path) -> dict[str, Any]:
        import openpyxl
        from openpyxl.utils import get_column_letter

        file_path = Path(file_path)
        _tmp_xlsx: Path | None = None
        if file_path.suffix.lower() == ".xls":
            from core.libreoffice_helper import convert_with_libreoffice

            file_path = convert_with_libreoffice(file_path, "xlsx")
            _tmp_xlsx = file_path

        try:
            return self._extract_styles_impl(file_path, openpyxl, get_column_letter)
        finally:
            if _tmp_xlsx and _tmp_xlsx.exists():
                _tmp_xlsx.unlink(missing_ok=True)

    def _extract_styles_impl(self, file_path: Path, openpyxl: Any, get_column_letter: Any) -> dict[str, Any]:
        styles: dict[str, Any] = {"document_level": {}}

        wb_data = openpyxl.load_workbook(str(file_path), data_only=True)
        try:
            wb_formula = openpyxl.load_workbook(str(file_path), data_only=False)
        except Exception:
            wb_formula = None

        styles["document_level"]["sheet_names"] = wb_data.sheetnames

        for sheet_name in wb_data.sheetnames:
            ws_data = wb_data[sheet_name]
            ws_formula = wb_formula[sheet_name] if wb_formula else None

            sheet_style: dict[str, Any] = {}

            # Column widths
            col_widths: dict[str, float] = {}
            for col_idx in range(ws_data.min_column or 1, (ws_data.max_column or 1) + 1):
                letter = get_column_letter(col_idx)
                dim = ws_data.column_dimensions.get(letter)
                if dim and dim.width:
                    col_widths[letter] = dim.width
            if col_widths:
                sheet_style["column_widths"] = col_widths

            # Row heights
            row_heights: dict[int, float] = {}
            for row_idx in range(ws_data.min_row or 1, (ws_data.max_row or 1) + 1):
                dim = ws_data.row_dimensions.get(row_idx)
                if dim and dim.height:
                    row_heights[row_idx] = dim.height
            if row_heights:
                sheet_style["row_heights"] = row_heights

            # Number formats per cell
            number_formats: dict[str, str] = {}
            for row in ws_data.iter_rows():
                for cell in row:
                    if cell.number_format and cell.number_format != "General":
                        number_formats[cell.coordinate] = cell.number_format
            if number_formats:
                sheet_style["number_formats"] = number_formats

            # Formulas from non-data_only workbook
            formulas: dict[str, str] = {}
            if ws_formula:
                for row in ws_formula.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            formulas[cell.coordinate] = cell.value
            if formulas:
                sheet_style["formulas"] = formulas

            # Merged cell ranges
            merged: list[str] = [str(mr) for mr in ws_data.merged_cells.ranges]
            if merged:
                sheet_style["merged_cells"] = merged

            # Freeze panes
            if ws_data.freeze_panes:
                sheet_style["freeze_panes"] = str(ws_data.freeze_panes)

            # Cell styles (sample first 100 cells)
            cell_styles: dict[str, dict] = {}
            cell_count = 0
            for row in ws_data.iter_rows():
                for cell in row:
                    if cell_count >= 100:
                        break
                    if cell.value is not None:
                        cs: dict[str, Any] = {}
                        if cell.font and cell.font.name:
                            cs["font"] = {
                                "name": cell.font.name,
                                "size": cell.font.size,
                                "bold": cell.font.bold,
                                "italic": cell.font.italic,
                            }
                        if cell.alignment and cell.alignment.horizontal != "general":
                            cs["alignment"] = {
                                "horizontal": cell.alignment.horizontal,
                                "vertical": cell.alignment.vertical,
                                "wrap_text": cell.alignment.wrap_text,
                            }
                        if cs:
                            cell_styles[cell.coordinate] = cs
                    cell_count += 1

            if cell_styles:
                sheet_style["cell_styles"] = cell_styles

            # Key by table content hash for sidecar lookup
            rows: list[list[str]] = []
            merged_map = self._build_merged_cells_map(ws_data)
            if ws_data.max_row and ws_data.max_column:
                for ri in range(ws_data.min_row, ws_data.max_row + 1):
                    row_data: list[str] = []
                    for ci in range(ws_data.min_column, ws_data.max_column + 1):
                        val = ws_data.cell(row=ri, column=ci).value
                        if val is None and (ri, ci) in merged_map:
                            val = merged_map[(ri, ci)]
                        row_data.append(self._format_cell_value(val))
                    rows.append(row_data)
                # Trim empty trailing rows
                while rows and all(c == "" for c in rows[-1]):
                    rows.pop()

            if rows:
                h = compute_content_hash(rows)
                styles[h] = sheet_style

            styles[f"sheet_{sheet_name}"] = sheet_style

        wb_data.close()
        if wb_formula:
            wb_formula.close()

        return styles

    @classmethod
    def supports_format(cls, extension: str) -> bool:
        return extension.lower().lstrip(".") in cls.EXTENSIONS
