"""Tests for formats/xlsx_handler.py — XLSX ingest, export, style extraction, round-trip."""

import tempfile
from pathlib import Path

import pytest

from core.document_model import ElementType
from formats.xlsx_handler import XlsxHandler


@pytest.fixture
def handler():
    return XlsxHandler()


# ── Ingest: simple.xlsx ─────────────────────────────────────────────────────

def test_ingest_returns_model(simple_xlsx, handler):
    model = handler.ingest(simple_xlsx)
    assert model is not None
    assert len(model.elements) > 0


def test_ingest_sheet_count(simple_xlsx, handler):
    """2 sheets → 2 H2 headings."""
    model = handler.ingest(simple_xlsx)
    headings = model.get_elements_by_type(ElementType.HEADING)
    h2s = [h for h in headings if h.level == 2]
    assert len(h2s) == 2


def test_ingest_sheet_names(simple_xlsx, handler):
    model = handler.ingest(simple_xlsx)
    headings = model.get_elements_by_type(ElementType.HEADING)
    names = [h.content for h in headings if h.level == 2]
    assert "Sales Data" in names
    assert "Summary" in names


def test_ingest_has_tables(simple_xlsx, handler):
    model = handler.ingest(simple_xlsx)
    tables = model.get_elements_by_type(ElementType.TABLE)
    assert len(tables) == 2  # One per sheet


def test_ingest_table_dimensions(simple_xlsx, handler):
    """Sheet 1 should have headers + 10 data rows + 1 formula row."""
    model = handler.ingest(simple_xlsx)
    tables = model.get_elements_by_type(ElementType.TABLE)
    sheet1_table = tables[0]
    rows = sheet1_table.content
    assert len(rows) >= 11  # header + 10 data rows (+ possible formula row)
    assert len(rows[0]) == 5  # 5 columns: Name, Region, Units, Price, Total


def test_ingest_cell_values(simple_xlsx, handler):
    """Spot-check specific cells."""
    model = handler.ingest(simple_xlsx)
    tables = model.get_elements_by_type(ElementType.TABLE)
    rows = tables[0].content
    assert rows[0][0] == "Name"  # Header
    assert rows[1][0] == "Alice"  # First data row


def test_ingest_merged_cells_unmerged(simple_xlsx, handler):
    """Sheet 2 has A2:A3 merged — both cells should have the value."""
    model = handler.ingest(simple_xlsx)
    tables = model.get_elements_by_type(ElementType.TABLE)
    sheet2_rows = tables[1].content
    # A2 and A3 should both be "Alpha" (merged and duplicated)
    assert sheet2_rows[1][0] == "Alpha"
    assert sheet2_rows[2][0] == "Alpha"


def test_ingest_metadata(simple_xlsx, handler):
    model = handler.ingest(simple_xlsx)
    assert model.metadata.source_format == "xlsx"
    assert model.metadata.page_count == 2


# ── Ingest: complex.xlsx ────────────────────────────────────────────────────

def test_ingest_complex_has_content(complex_xlsx, handler):
    model = handler.ingest(complex_xlsx)
    tables = model.get_elements_by_type(ElementType.TABLE)
    assert len(tables) >= 1
    rows = tables[0].content
    assert len(rows) >= 2  # header + data


# ── Export ───────────────────────────────────────────────────────────────────

def test_export_creates_file(simple_xlsx, handler, tmp_path):
    model = handler.ingest(simple_xlsx)
    output = tmp_path / "output.xlsx"
    handler.export(model, output)
    assert output.exists()
    assert output.stat().st_size > 0


def test_export_sheet_count(simple_xlsx, handler, tmp_path):
    import openpyxl

    model = handler.ingest(simple_xlsx)
    output = tmp_path / "output.xlsx"
    handler.export(model, output)

    wb = openpyxl.load_workbook(str(output))
    assert len(wb.sheetnames) == 2
    wb.close()


def test_export_sheet_names(simple_xlsx, handler, tmp_path):
    import openpyxl

    model = handler.ingest(simple_xlsx)
    output = tmp_path / "output.xlsx"
    handler.export(model, output)

    wb = openpyxl.load_workbook(str(output))
    assert "Sales Data" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    wb.close()


def test_export_cell_values_match(simple_xlsx, handler, tmp_path):
    import openpyxl

    model = handler.ingest(simple_xlsx)
    output = tmp_path / "output.xlsx"
    handler.export(model, output)

    wb = openpyxl.load_workbook(str(output), data_only=True)
    ws = wb["Sales Data"]
    assert ws.cell(row=1, column=1).value == "Name"
    assert ws.cell(row=2, column=1).value == "Alice"
    wb.close()


def test_export_with_sidecar_formulas(simple_xlsx, handler, tmp_path):
    """Sidecar Tier 2: formulas should be restored from sidecar."""
    import openpyxl

    model = handler.ingest(simple_xlsx)
    styles = handler.extract_styles(simple_xlsx)

    # Build sidecar
    sidecar = {
        "document_level": styles.get("document_level", {}),
        "elements": {k: v for k, v in styles.items() if k not in ("document_level",)},
    }

    output = tmp_path / "styled.xlsx"
    handler.export(model, output, sidecar=sidecar)

    # Verify formula is restored
    wb = openpyxl.load_workbook(str(output), data_only=False)
    ws = wb.active
    # Check if any formula cells exist
    has_formula = False
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                has_formula = True
                break
    wb.close()
    # Formulas should be restored from sidecar
    assert has_formula


# ── Style extraction ─────────────────────────────────────────────────────────

def test_extract_styles_document_level(simple_xlsx, handler):
    styles = handler.extract_styles(simple_xlsx)
    assert "document_level" in styles
    assert "sheet_names" in styles["document_level"]


def test_extract_styles_has_formulas(simple_xlsx, handler):
    """Style extraction should capture formulas."""
    styles = handler.extract_styles(simple_xlsx)
    # Check any sheet-level style has formulas
    has_formulas = False
    for key, val in styles.items():
        if isinstance(val, dict) and "formulas" in val:
            has_formulas = True
            break
    assert has_formulas


def test_extract_styles_column_widths(simple_xlsx, handler):
    styles = handler.extract_styles(simple_xlsx)
    has_widths = False
    for key, val in styles.items():
        if isinstance(val, dict) and "column_widths" in val:
            has_widths = True
            break
    assert has_widths


# ── Round-trip ───────────────────────────────────────────────────────────────

def test_roundtrip_sheet_count(simple_xlsx, handler, tmp_path):
    """XLSX → DocumentModel → XLSX: 2 sheets survive."""
    import openpyxl

    model = handler.ingest(simple_xlsx)
    output = tmp_path / "roundtrip.xlsx"
    handler.export(model, output)

    wb = openpyxl.load_workbook(str(output))
    assert len(wb.sheetnames) == 2
    wb.close()


def test_roundtrip_values_match(simple_xlsx, handler, tmp_path):
    import openpyxl

    model = handler.ingest(simple_xlsx)
    output = tmp_path / "roundtrip.xlsx"
    handler.export(model, output)

    wb = openpyxl.load_workbook(str(output), data_only=True)
    ws = wb["Sales Data"]
    assert ws.cell(row=1, column=1).value == "Name"
    assert ws.cell(row=2, column=1).value == "Alice"
    assert ws.cell(row=2, column=3).value == 10  # Units
    wb.close()


def test_roundtrip_row_column_counts(simple_xlsx, handler, tmp_path):
    """Row and column counts should match after round-trip."""
    model = handler.ingest(simple_xlsx)
    tables = model.get_elements_by_type(ElementType.TABLE)
    orig_rows = len(tables[0].content)
    orig_cols = len(tables[0].content[0])

    output = tmp_path / "roundtrip.xlsx"
    handler.export(model, output)

    model2 = handler.ingest(output)
    tables2 = model2.get_elements_by_type(ElementType.TABLE)
    assert len(tables2[0].content) == orig_rows
    assert len(tables2[0].content[0]) == orig_cols


# ── supports_format ──────────────────────────────────────────────────────────

def test_supports_xlsx():
    assert XlsxHandler.supports_format("xlsx")
    assert XlsxHandler.supports_format(".xlsx")
    assert not XlsxHandler.supports_format("csv")
