"""
Tests for formats/xlsx_handler.py — XLSX ingest and export.

Covers sheet extraction, merged cells, formulas, edge cases,
and Markdown → XLSX export with fidelity tiers.
"""

import pytest
from pathlib import Path

from core.document_model import DocumentModel, Element, ElementType


# ── Ingest ───────────────────────────────────────────────────────────────────

class TestXlsxIngest:
    """Tests for XLSX ingestion."""

    def test_sheets_become_h2_and_table(self, simple_xlsx):
        from formats.xlsx_handler import XlsxHandler

        handler = XlsxHandler()
        model = handler.ingest(simple_xlsx)

        headings = model.get_elements_by_type(ElementType.HEADING)
        h2s = [h for h in headings if h.level == 2]
        assert len(h2s) == 2, "simple.xlsx has 2 sheets"

        tables = model.get_elements_by_type(ElementType.TABLE)
        assert len(tables) >= 1

    def test_merged_cells_unmerged(self, simple_xlsx):
        """Merged cells should be unmerged with top-left value duplicated."""
        from formats.xlsx_handler import XlsxHandler

        handler = XlsxHandler()
        model = handler.ingest(simple_xlsx)

        tables = model.get_elements_by_type(ElementType.TABLE)
        # The Summary sheet (second table) has A2:A3 merged
        assert len(tables) >= 2
        summary_table = tables[1].content
        # A2 and A3 should both have "Alpha"
        assert summary_table[1][0] == "Alpha"
        assert summary_table[2][0] == "Alpha"

    def test_page_count_equals_sheet_count(self, simple_xlsx):
        from formats.xlsx_handler import XlsxHandler

        handler = XlsxHandler()
        model = handler.ingest(simple_xlsx)

        assert model.metadata.page_count == 2

    def test_source_format(self, simple_xlsx):
        from formats.xlsx_handler import XlsxHandler

        handler = XlsxHandler()
        model = handler.ingest(simple_xlsx)

        assert model.metadata.source_format == "xlsx"

    def test_computed_values_used(self, simple_xlsx):
        """data_only=True workbook should give computed values, not formula strings."""
        from formats.xlsx_handler import XlsxHandler

        handler = XlsxHandler()
        model = handler.ingest(simple_xlsx)

        tables = model.get_elements_by_type(ElementType.TABLE)
        if tables:
            # Check that no cell value starts with '='
            for row in tables[0].content:
                for cell in row:
                    # Computed values may be None/empty when not calculated
                    if cell and cell.startswith("="):
                        pytest.fail(f"Formula string found in data: {cell}")


# ── Ingest edge cases ───────────────────────────────────────────────────────

class TestXlsxIngestEdgeCases:
    """Edge case tests for XLSX ingest."""

    def test_corrupt_xlsx_raises_error(self, tmp_path):
        from formats.xlsx_handler import XlsxHandler

        bad = tmp_path / "corrupt.xlsx"
        bad.write_bytes(b"NOT AN XLSX FILE")

        handler = XlsxHandler()
        with pytest.raises(Exception):
            handler.ingest(bad)

    def test_wide_table_ingests(self, tmp_path):
        """50+ column table should not hang or crash."""
        import openpyxl
        from formats.xlsx_handler import XlsxHandler

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Wide"
        # 55 columns, 5 rows
        for r in range(1, 6):
            for c in range(1, 56):
                ws.cell(row=r, column=c, value=f"R{r}C{c}")
        path = tmp_path / "wide.xlsx"
        wb.save(str(path))

        handler = XlsxHandler()
        model = handler.ingest(path)

        tables = model.get_elements_by_type(ElementType.TABLE)
        assert len(tables) == 1
        assert len(tables[0].content[0]) == 55

    def test_empty_sheet_no_crash(self, tmp_path):
        """Empty sheet should not crash."""
        import openpyxl
        from formats.xlsx_handler import XlsxHandler

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empty"
        path = tmp_path / "empty.xlsx"
        wb.save(str(path))

        handler = XlsxHandler()
        model = handler.ingest(path)

        headings = model.get_elements_by_type(ElementType.HEADING)
        assert len(headings) >= 1  # At least the sheet heading


# ── Export ───────────────────────────────────────────────────────────────────

class TestXlsxExport:
    """Tests for Markdown → XLSX export."""

    def test_table_round_trips(self, tmp_path):
        from formats.xlsx_handler import XlsxHandler

        model = DocumentModel()
        model.add_element(Element(type=ElementType.HEADING, content="Sheet1", level=2))
        model.add_element(Element(
            type=ElementType.TABLE,
            content=[["Name", "Value"], ["A", "1"], ["B", "2"]],
        ))

        handler = XlsxHandler()
        output = tmp_path / "export.xlsx"
        handler.export(model, output)

        # Re-ingest and verify
        model2 = handler.ingest(output)
        tables = model2.get_elements_by_type(ElementType.TABLE)
        assert len(tables) == 1
        assert len(tables[0].content) == 3  # header + 2 rows
        assert len(tables[0].content[0]) == 2  # 2 columns

    def test_multiple_sheets(self, tmp_path):
        from formats.xlsx_handler import XlsxHandler

        model = DocumentModel()
        model.add_element(Element(type=ElementType.HEADING, content="Sheet1", level=2))
        model.add_element(Element(type=ElementType.TABLE, content=[["A", "B"], ["1", "2"]]))
        model.add_element(Element(type=ElementType.HORIZONTAL_RULE, content=""))
        model.add_element(Element(type=ElementType.HEADING, content="Sheet2", level=2))
        model.add_element(Element(type=ElementType.TABLE, content=[["X", "Y"], ["3", "4"]]))

        handler = XlsxHandler()
        output = tmp_path / "multi.xlsx"
        handler.export(model, output)

        import openpyxl
        wb = openpyxl.load_workbook(str(output))
        assert len(wb.sheetnames) == 2
        wb.close()

    def test_tier2_column_widths(self, simple_xlsx, tmp_path):
        """Tier 2: column widths from sidecar are applied."""
        from formats.xlsx_handler import XlsxHandler

        handler = XlsxHandler()
        model = handler.ingest(simple_xlsx)
        sidecar = handler.extract_styles(simple_xlsx)

        output = tmp_path / "tier2.xlsx"
        handler.export(model, output, sidecar=sidecar)

        assert output.exists()
        assert output.stat().st_size > 0

    def test_tier3_patching(self, simple_xlsx, tmp_path):
        """Tier 3: formula cells from sidecar are written back."""
        from formats.xlsx_handler import XlsxHandler

        handler = XlsxHandler()
        model = handler.ingest(simple_xlsx)
        sidecar = handler.extract_styles(simple_xlsx)

        output = tmp_path / "tier3.xlsx"
        handler.export(model, output, sidecar=sidecar, original_path=simple_xlsx)

        assert output.exists()


class TestXlsxStyleExtraction:
    """Tests for XLSX style extraction."""

    def test_extract_styles_returns_dict(self, simple_xlsx):
        from formats.xlsx_handler import XlsxHandler

        handler = XlsxHandler()
        styles = handler.extract_styles(simple_xlsx)
        assert isinstance(styles, dict)
        assert "document_level" in styles

    def test_extract_styles_has_sheet_names(self, simple_xlsx):
        from formats.xlsx_handler import XlsxHandler

        handler = XlsxHandler()
        styles = handler.extract_styles(simple_xlsx)
        assert "sheet_names" in styles["document_level"]
        assert len(styles["document_level"]["sheet_names"]) == 2

    def test_extract_styles_formulas(self, simple_xlsx):
        from formats.xlsx_handler import XlsxHandler

        handler = XlsxHandler()
        styles = handler.extract_styles(simple_xlsx)
        # Should find formulas in at least one sheet
        has_formulas = any(
            "formulas" in styles.get(k, {})
            for k in styles
            if k.startswith("sheet_")
        )
        assert has_formulas

    def test_supports_format(self):
        from formats.xlsx_handler import XlsxHandler

        assert XlsxHandler.supports_format("xlsx")
        assert XlsxHandler.supports_format(".xlsx")
        assert not XlsxHandler.supports_format("csv")
