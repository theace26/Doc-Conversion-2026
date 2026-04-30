"""Generate phase_model_plan.xlsx — Claude Code model + effort recommendations
per phase for the UX Overhaul and Active Operations Registry (System Unifier).

Effort units = Claude Code session-hours, t-shirt sized:
    XS ≈ <2h    S ≈ 2-4h    M ≈ 4-8h    L ≈ 8-16h    XL ≈ 16-32h

Reviewer effort is for one focused review pass against the spec/plan, not for
back-and-forth iteration. Real cost = sessions × model rate.

Run:  python3 build_phase_model_plan.py
Output: phase_model_plan.xlsx (same dir)
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "phase_model_plan.xlsx"

HEADERS = [
    "Phase",
    "Source plan file",
    "Implementer model",
    "Implementer effort",
    "Reviewer model",
    "Reviewer effort",
    "Reasoning",
]

UX_OVERHAUL = [
    (
        "Plan 1A — Foundation Setup",
        "2026-04-28-ux-overhaul-foundation-setup.md",
        "Sonnet",   "M (4-8h)",
        "Opus",     "S (2-3h)",
        "Foundation work: aiosqlite migration for the user_preferences table, JWT role-claim "
        "parsing, design-token CSS, prefs server module + endpoints. Schema and auth-claim "
        "mistakes are expensive to undo once data lands, so spend once on Opus review even "
        "though Sonnet handles the implementation comfortably. Effort is medium because the "
        "scope is wide but each piece is well-defined; reviewer focuses on the migration and "
        "JWT slice (the irreversible bits), not the CSS tokens.",
    ),
    (
        "Plan 1B — Static Chrome",
        "2026-04-28-ux-overhaul-static-chrome.md",
        "Haiku",    "S (2-4h)",
        "Haiku",    "XS (<1h)",
        "Four pure-presentational vanilla-JS components (top-nav, version-chip, avatar, "
        "layout-icon), each ≤100 LOC, no state, no async, no server interaction. Pattern is "
        "repetitive and the safe-DOM rules are explicit. Cheapest tier on both sides; the "
        "review just spot-checks for innerHTML and pattern adherence — fits in one short pass.",
    ),
    (
        "Plan 1C — Stateful Chrome",
        "2026-04-28-ux-overhaul-stateful-chrome.md",
        "Sonnet",   "M (4-8h)",
        "Sonnet",   "S (2-3h)",
        "Prefs client (localStorage cache + 500ms debounced PUT), telemetry helper (fire-and-"
        "forget), avatar/layout popovers with click-outside + Escape. Subtle async + event-"
        "lifecycle but well-bounded. Reviewer needs to follow the popover state machine + "
        "debounce semantics, hence S not XS.",
    ),
    (
        "Plan 2A — Document Card + Density Modes",
        "2026-04-28-ux-overhaul-document-card.md",
        "Haiku",    "S (2-4h)",
        "Sonnet",   "XS (<1h)",
        "Presentational card component with three density modes driven by CSS variables and "
        "MFPrefs. Implementation is mechanical given the explicit mockups, but bumping reviewer "
        "to Sonnet because density toggle interacts with the prefs system landed in 1C — worth "
        "verifying the binding is correct, not just the visuals. Review fits in a single short "
        "pass since the surface area is small.",
    ),
    (
        "Plan 2B — Card Interactions",
        "2026-04-28-ux-overhaul-card-interactions.md",
        "Sonnet",   "L (8-16h)",
        "Sonnet",   "M (3-5h)",
        "Hover preview popover, right-click context menu, multi-select observable state, bulk "
        "action bar, folder-browse page wrapping it all. Multiple interaction systems coexisting "
        "without listener leaks is the failure mode. This is the biggest UX-overhaul phase — "
        "reviewer needs M because there are 4 distinct interaction surfaces to walk through, "
        "not one.",
    ),
    (
        "Plan 3 — Search-as-Home Page",
        "2026-04-28-ux-overhaul-search-as-home.md",
        "Sonnet",   "M (4-8h)",
        "Sonnet",   "S (2-3h)",
        "Feature-flagged replacement of static/index.html with three layout modes (Maximal / "
        "Recent / Minimal). Both flag-on and flag-off paths must work — flag-off keeps legacy "
        "Convert page live, so rollback is cheap. Reviewer's main job is verifying the flag "
        "branching and that legacy still renders.",
    ),
    (
        "Plan 4 — IA Shift",
        "2026-04-28-ux-overhaul-ia-shift.md",
        "Sonnet",   "M (4-8h)",
        "Opus",     "S (2-3h)",
        "New /api/me endpoint, real role binding via UnionCore JWT (replaces the hardcoded "
        "role='admin' placeholder), Activity dashboard gated by core.auth.Role. Auth surface — "
        "role-gating bugs are security issues. Opus reviewer once to verify the gate is enforced "
        "on every protected endpoint, not just the obvious ones; effort is small because the new "
        "endpoint surface is narrow.",
    ),
]

SYSTEM_UNIFIER = [
    (
        "Phase 0 — Pre-flight reconnaissance (Tasks 0.1–0.7)",
        "2026-04-28-active-operations-registry.md",
        "Haiku",    "S (2-4h)",
        "Haiku",    "XS (<1h)",
        "Read-only discovery: trace existing pipeline.py, scan_coordinator.py, bulk_worker.py, "
        "lifecycle_scanner.py to confirm assumptions about cancel signals and lifecycle hooks. "
        "Output is notes, not code. Cheapest tier on both sides; review is just a sanity-check "
        "that the recon notes match the plan's stated assumptions.",
    ),
    (
        "Phase 1 — Registry + DB foundation (Tasks 1–10)",
        "2026-04-28-active-operations-registry.md",
        "Opus",     "L (8-16h)",
        "Opus",     "M (3-5h)",
        "Load-bearing concurrency primitive: new core/active_ops.py (in-memory dict + write-"
        "through to a new active_operations SQLite table), lifespan hooks, scheduler "
        "integration, singleton lifecycle. Every later phase depends on this — race conditions "
        "or write-through gaps cascade across all six worker subsystems. Spend tokens here on "
        "both sides; this is the genuinely highest-risk slice of the whole plan and the only "
        "place where Opus/Opus is unconditionally the right call.",
    ),
    (
        "Phase 2 — HTTP API (Tasks 11–13)",
        "2026-04-28-active-operations-registry.md",
        "Sonnet",   "S (2-4h)",
        "Sonnet",   "XS (<1h)",
        "Single GET /api/active-ops endpoint feeding three frontend surfaces. Standard FastAPI "
        "pattern with a strict response schema. Review is a quick contract check against Phase "
        "4's three consumers; payload is small and readable in one screen.",
    ),
    (
        "Phase 3 — Worker retrofits (Tasks 14–23)",
        "2026-04-28-active-operations-registry.md",
        "Opus",     "XL (16-32h)",
        "Sonnet",   "L (5-8h)",
        "Touches six worker modules (pipeline, scan_coordinator, trash, analysis, lifecycle_"
        "scanner, bulk_worker) to register/update/finish operations and bridge cancel signals "
        "to each subsystem's native mechanism. Largest phase by far. The structural traps — "
        "double-register, missed cleanup on exception paths, cancel hook never firing — need "
        "Opus to catch on the implementer side. Reviewer at Sonnet has 6 modules × cancel "
        "paths to walk through, hence L not M.",
    ),
    (
        "Phase 4 — Frontend (Tasks 24–34)",
        "2026-04-28-active-operations-registry.md",
        "Sonnet",   "L (8-16h)",
        "Haiku",    "S (2-3h)",
        "Three presentational surfaces (sticky banner, per-page inline widget, Status hub) all "
        "fed by the /api/active-ops payload from Phase 2. 11 tasks across the three surfaces. "
        "Once that contract is stable the work is mechanical DOM construction + polling. Haiku "
        "reviewer is fine — visual diff against mockups is exactly what cheap models do well — "
        "but they need S not XS because there are three surfaces to compare.",
    ),
    (
        "Phase 5 — Cleanup + documentation (Tasks 35–46)",
        "2026-04-28-active-operations-registry.md",
        "Sonnet",   "M (4-8h)",
        "Sonnet",   "S (2-3h)",
        "Removing dead per-subsystem progress code that the registry now subsumes, plus "
        "CLAUDE.md / version-history.md / whats-new.md updates. 12 tasks. Deletions are "
        "irreversible — judgement about what's truly subsumed vs. what still has unique callers "
        "needs Sonnet, not Haiku. Reviewer focus is on the deletions, not the doc edits.",
    ),
]


def write_sheet(ws, rows):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    model_fills = {
        "Haiku":  PatternFill("solid", fgColor="C6EFCE"),
        "Sonnet": PatternFill("solid", fgColor="FFEB9C"),
        "Opus":   PatternFill("solid", fgColor="FFC7CE"),
    }
    effort_fills = {
        "XS": PatternFill("solid", fgColor="EAF3FB"),
        "S":  PatternFill("solid", fgColor="D9EAD3"),
        "M":  PatternFill("solid", fgColor="FFF2CC"),
        "L":  PatternFill("solid", fgColor="FCE5CD"),
        "XL": PatternFill("solid", fgColor="F4CCCC"),
    }
    wrap = Alignment(wrap_text=True, vertical="top")
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = centre

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = wrap
            if c_idx in (3, 5):  # model columns
                if val in model_fills:
                    c.fill = model_fills[val]
                    c.font = Font(bold=True)
                    c.alignment = centre
            elif c_idx in (4, 6):  # effort columns
                size = str(val).split()[0]
                if size in effort_fills:
                    c.fill = effort_fills[size]
                    c.alignment = centre

    widths = {1: 44, 2: 48, 3: 13, 4: 13, 5: 13, 6: 13, 7: 88}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28
    for r in range(2, 2 + len(rows)):
        ws.row_dimensions[r].height = 120
    ws.freeze_panes = "A2"


def write_legend(ws):
    ws["A1"] = "Model legend & cost guidance"
    ws["A1"].font = Font(bold=True, size=14)
    rows = [
        ("Model", "Use for", "Approx cost vs Opus"),
        ("Haiku",  "Pattern-driven scaffolding, presentational vanilla-JS components, doc edits, read-only discovery, visual diff review.", "≈ 1/15"),
        ("Sonnet", "Most feature work: FastAPI endpoints, vanilla-JS with state, multi-system interactions, judgement-call cleanup.",       "≈ 1/5"),
        ("Opus",   "Concurrency primitives, schema/migration design, auth surface, cancel-propagation across subsystems.",                 "1×"),
        ("", "", ""),
        ("Effort size", "Approx Claude Code session-hours", ""),
        ("XS", "<2h — single short pass; cosmetic edits, narrow contract checks.", ""),
        ("S",  "2-4h — one focused session.",                                       ""),
        ("M",  "4-8h — half-day to a day; multi-file feature.",                     ""),
        ("L",  "8-16h — one to two days; multiple interacting components.",         ""),
        ("XL", "16-32h — multi-day; touches several subsystems (e.g. Phase 3 worker retrofits).", ""),
        ("", "", ""),
        ("Heuristic", "Default both implementer and reviewer to Sonnet/M. Justify every Opus upgrade (irreversibility, security, concurrency) and every Haiku downgrade (mechanical, well-spec'd, low blast radius). Reviewer effort runs ~25-35% of implementer effort for a single focused pass — bigger if the implementer touched many subsystems.", ""),
    ]
    for r, row in enumerate(rows, start=3):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r == 3 or (val and r in (8,)):
                cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 100
    ws.column_dimensions["C"].width = 22
    for r in range(4, 4 + len(rows) - 1):
        ws.row_dimensions[r].height = 38


def main():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "UX Overhaul"
    write_sheet(ws1, UX_OVERHAUL)

    ws2 = wb.create_sheet("System Unifier")
    write_sheet(ws2, SYSTEM_UNIFIER)

    ws3 = wb.create_sheet("Legend")
    write_legend(ws3)

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
